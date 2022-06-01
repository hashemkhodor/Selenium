# -*- coding: utf-8 -*-
"""
Created on Sat May 21 17:28:36 2022

@author: Hashem
"""
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import constants
import time
from openpyxl import Workbook
from Product import Product
class AmazonBot(object):
    def __init__(self,path):
        self.path = path
        self.driver= None
        self.currentproduct=None
        self.handles=[]
        self.products=[]
        self.driver= webdriver.Chrome(self.path)
        self.driver.execute_script("window.open('');")
        self.handles=self.driver.window_handles
        self.wait= WebDriverWait(self.driver,10)
        self.driver.switch_to.window(self.handles[0])
    
    def __click(self,button):
        #assert type(button)==webdriver.WebElement ,"Bad Input"
        while True:
            try:
                webdriver.ActionChains(self.driver).move_to_element(button).click().perform()
                time.sleep(3)
                return
            except:
                continue
    def __getproductlinks(self,pages=1):
        if pages==1:
            Links= self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,'a[class="a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal"]')))
            links=[link.get_attribute("href") for link in Links]
        else:
            links=[]
            for i in range(pages):
                time.sleep(2)
                Links= self.find_elements_by_css_selector('a[class="a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal"]')
                for link in Links:
                    links.append(link.get_attribute("href"))
                nextpage= self.find_elements_by_css_selector('div[role="navigation"] a[aria-label*="next page"]')
                # assert nextpage , "Maximum number of pages reached" +str(i+1)
                if nextpage:
                    self.driver.get(nextpage[0].get_attribute("href"))
                else:
                    print("The maximum number of pages is: ", i+1)
                
                
        return links
    def getproductdelivery(self,start):
        delivery1=self.driver.find_elements_by_xpath(start+'//div[contains(@id,"mir-layout-DELIVERY_BLOCK")]/span[not(contains(@data-csa-c-type,"element"))]')
        delivery2=self.driver.find_elements_by_xpath(start+'//div[contains(@id,"mir-layout-DELIVERY_BLOCK")]/span[contains(@data-csa-c-type,"element")]')
        d=[]
        if delivery1:
            for deliv in delivery1:
                d.append(deliv.text)
        elif delivery2:
            for deliv in delivery2:
                d.append((deliv.get_attribute('data-csa-c-delivery-price'),deliv.get_attribute('data-csa-c-delivery-time')))
            
        return d
    
    def getsellerfrom(self,e):
        xpath='(//div[contains(@id,"shipsFrom")])['+str(e)+'] //div[contains(@class,"col-right")] //span[contains(@class,"a-size-small ")]'
        shipsfrom=self.wait.until(EC.presence_of_all_elements_located((By.XPATH,xpath)))
        seller=shipsfrom[0].get_attribute("textContent")
        country= shipsfrom[1].get_attribute("textContent")
        return seller,country
    
    def __buyingoptions(self):
       buyingoptions=self.driver.find_elements_by_css_selector('span[id="buybox-see-all-buying-choices"] [title="See All Buying Options"]')
       if buyingoptions:
           alloptions=[]
           self.__click(buyingoptions[0])
           totaloptions= self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR,'div[id="aod-message-component"]>span[id="aod-filter-offer-count-string"]')))
           totaloptions= totaloptions.get_attribute("textContent").split(" ")
           totaloptions=int(totaloptions[0])
           for i in range(1,totaloptions+1):
               option=[]
               d=[]
               price= self.wait.until(EC.presence_of_element_located((By.XPATH,'(//div[@id="aod-offer"]//span[@class="a-offscreen"])['+str(i)+']')))
               d=self.getproductdelivery('(//div[@id="aod-offer"])['+str(i)+']')
              # d2=self.getsellerfrom(i)
               option= (price.get_attribute("textContent"),d)
               print(option,end=" ")
               alloptions.append(option)
           
           self.driver.back()
           return alloptions
       else:
           return []
    
    def getproductname(self):
        ## To be added: update the constants
        time.sleep(2)
        try:
            name=self.wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,constants.productname)))
            return name[0].get_attribute("textContent")
        except:
            return self.getproductname()
        
       # self.wait.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR,constants.productname),name[0].get_attribute("textContent")))
        #return name[0].get_attribute("textContent")
        
    
    def getproductpriceanddelivery(self):
        m=[]
        prices=self.driver.find_elements_by_css_selector(constants.price)
       
        if len(prices)!=0:
            p=prices[0].get_attribute("textContent")
            #m.append(p)
            d=self.getproductdelivery('')
            
            m=(p,d)    
        A=self.__buyingoptions()
        if len(m)!=0:
            A.append(m)
        if A :
            return A
        else:
            price= self.driver.find_elements_by_css_selector('div[id^="out"] span[class*="price"]')
            if price:
                return (price[0].get_attribute("textContent"),"Unavailable")
            
    def getoptions(self): 
        var1=self.driver.find_elements_by_xpath('//div[contains(@id,"inline-twister-row")] ')
        var2= self.driver.find_elements_by_xpath('//div[contains(@id,"variation")]')
        typess={}
        var=""
        if var1:
            var=1
            totaltypes=len(var1)
            for i in range(0,totaltypes):
                #print(var1[i].get_attribute("id").split("_")[0].split("-")[-1],end=" _ ")
                ## here We're gettind the variations names. Example( Beat headphones differ from their colors)
                typess[var1[i].get_attribute("id").split("_")[0].split("-")[-1]]=True  
        elif var2:
            var=2
            totaltypes=len(var2)
            for i in range(0,totaltypes):
               # print(var2[i].get_attribute("id").split("_")[1],end=" , ")
               ## here We're gettind the variations names. Example( Beat headphones differ from their colors)
                typess[var2[i].get_attribute("id").split("_")[1]]=1    
        return list(typess.keys()),var
    
    def get_path_option(self,option,var):
        if var==1:
            return '//div[contains(@id,"inline-twister-row")]//li[@data-asin]//span[contains(@id,"' + option +'")][not(contains(@class,"unavailable"))]/span/ input | button'
        elif var==2:
            #//div[contains(@id,"variation")]// li[contains(@id,"color")]// button | input
            return '//div[contains(@id,"variation")]//li[contains(@id,"'  +option+'")] // button | input'
    
    def scrapepage_updated(self,link,options,var):
       good , critical, rating= self.get_product_reviews_and_rating()
       if len(options) ==0:
           name=self.getproductname()
           duos=self.getproductpriceanddelivery()
           overview,features,details= self.get_product_overview_and_features()
           good , critical, rating= self.get_product_reviews_and_rating(link)
           product=Product(name=name, deliveryoptions=duos,link=link,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
           product.deliverybreakdown()
           self.products.append(product)
        
       elif len(options) ==1:
            xpath=self.get_path_option(options[0],var)
            option1= self.driver.find_elements_by_xpath(xpath)
            for op1 in option1:
                self.__click(op1)
                time.sleep(1.5)
                name=self.getproductname()
                duos=self.getproductpriceanddelivery()
                overview,features,details= self.get_product_overview_and_features()
                product=Product(name=name,link=link,keyword=self.currentproduct,deliveryoptions=duos,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                product.deliverybreakdown()
                self.products.append(product)  
        
       elif len(options)==2:
            xpath=self.get_path_option(options[0],var)
            option1= self.driver.find_elements_by_xpath(xpath)
            for op1 in option1:
                self.__click(op1)
                time.sleep(1.5)
                option2=self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[1],var))))
                for op2 in option2:
                    self.__click(op2)
                    name=self.getproductname()
                    duos=self.getproductpriceanddelivery()
                    overview,features,details= self.get_product_overview_and_features()
                    product=Product(name=name,link=link,keyword=self.currentproduct,deliveryoptions=duos,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                    product.deliverybreakdown()
                    self.products.append(product)  
        
       elif len(options)==3:
            xpath=self.get_path_option(options[0],var)
            option1= self.driver.find_elements_by_xpath(xpath)
            for op1 in option1:
                self.__click(op1)
                option2=self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[1],var))))
                for op2 in option2:
                    self.__click(op2)
                    option3= self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[2],var))))
                    for op3 in option3:
                        self.__click(op3)
                        time.sleep(1.5)
                        name=self.getproductname()
                        duos=self.getproductpriceanddelivery()
                        overview,features,details= self.get_product_overview_and_features()
                        product=Product(name=name,link=link,keyword=self.currentproduct,deliveryoptions=duos,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                        product.deliverybreakdown()
                        self.products.append(product) 
        
       elif len(options)==4:
            xpath=self.get_path_option(options[0],var)
            option1= self.driver.find_elements_by_xpath(xpath)
            for op1 in option1:
                self.__click(op1)
                option2=self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[1],var))))
                for op2 in option2:
                    self.__click(op2)
                    option3= self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[2],var))))
                    for op3 in option3:
                        self.__click(op3)
                        option4= self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[3],var))))
                        for op4 in option4:
                            self.__click(op4)
                            time.sleep(1.5)
                            name=self.getproductname()
                            duos=self.getproductpriceanddelivery()
                            overview,features,details= self.get_product_overview_and_features()
                            product=Product(name=name,link=link,keyword=self.currentproduct,deliveryoptions=duos,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                            product.deliverybreakdown()
                            self.products.append(product) 
    
    
    def scrapepage(self,link):
        self.driver.get(link)
        options,var= self.getoptions()
        D={}
        for option in options:
            if self.driver.find_elements_by_xpath(self.get_path_option(option,var)):
                D[option]=True
        self.scrapepage_updated(link,list(D.keys()),var)
        return
    
    def scrapeproduct(self,product,pages=2):
        self.currentproduct=product
        self.driver.get('https://www.amazon.com/')
        searchBar=self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,constants.search_bar)))
        searchBar.send_keys(product)
        searchBar.send_keys(Keys.RETURN)
        links= self.__getproductlinks(pages)
       
        for link in links:
            self.scrapepage(link)
            
        return
    
    def get_product_overview_and_features(self):
        rows= self.driver.find_elements_by_css_selector('div[id*="Overview"] tr')
        bullets= self.driver.find_elements_by_css_selector('div[id*="featurebullets"] ul li')
        description= self.driver.find_elements_by_css_selector('div[id="productDescription"] p')
        details1=self.driver.find_elements_by_css_selector('div[id*="detail"] ul li')
        details2= self.driver.find_elements_by_xpath('//div[@id="prodDetails"]//tr /th /following-sibling:: td[contains(@class,"prodDetAttrValue")]')
        
        overview=""
        bultext=""
        detail1=""
        detail2=""
        for i in range(1,len(rows)+1):
            cat= self.driver.find_elements_by_css_selector('div[id*="Overview"] tr:nth-child('+str(i)+') td:nth-child(1)')
            info= self.driver.find_elements_by_css_selector('div[id*="Overview"] tr:nth-child('+str(i)+') td:nth-child(2)')
            if cat and info:
                
                overview+= cat[0].get_attribute("textContent") + ":" + info[0].get_attribute("textContent") 
            if i != len(rows):
                overview+=" \n"
        for j in range(3,len(bullets)+3):
            
            bullet= self.driver.find_elements_by_css_selector('div[id*="featurebullets"] ul li:nth-child('+str(j)+')')
            if bullet:
                bultext += bullet[0].get_attribute("textContent")
                if j != len(bullets)-1 and bultext[-1:] != "\n":
                    bultext+= " \n"
        for desc in description:
            overview+= desc.get_attribute("textContent")
        
        for k in range(1,len(details1)+1):
            det=self.driver.find_elements_by_xpath('//div[@id="detailBullets_feature_div"]// li['+str(k)+']/span/span') 
            
            if det:
                detail1+= det[0].get_attribute("textContent") +" : " + det[1].get_attribute("textContent") # detaill[0] is the domain and detaill[1] is the value
                if k != len(details1):
                    detail1+=" \n"
        for k in range(1,len(details2)+1):
            detaill=self.driver.find_elements_by_xpath('(//div[@id="prodDetails"]//tr /th /following-sibling:: td[contains(@class,"prodDetAttrValue")])['+str(k)+']')
            attribute= self.driver.find_elements_by_xpath( '(//div[@id="prodDetails"]//tr/ td[contains(@class,"prodDetAttrValue")] )['+str(k)+']/preceding-sibling::th ')
            if detaill and attribute:
                detail1+= attribute[0].get_attribute("textContent")+ ":" + detaill[0].get_attribute("textContent")
                if k != len(details1):
                    detail1+=" \n"
            
        return overview , bultext , detail1+detail2
    def extract_reviews(self,reviews):
      
        link= reviews[0].get_attribute("href")
        self.driver.switch_to.window(self.handles[-1])
        g_reviews=""
        c_reviews=""
        if not reviews:
            self.driver.switch_to.window(self.handles[0])
            return ('','')
        
        
        self.driver.get(link)
        goodreviews=self.driver.find_elements_by_xpath('//a[contains(@data-reviews-state-param,"positive")]')
        criticalreviews=self.driver.find_elements_by_xpath('//a[contains(@data-reviews-state-param,"critical") ] ')
        if goodreviews:
            link1= goodreviews[0].get_attribute("href")
        if criticalreviews:
            link2= criticalreviews[0].get_attribute("href")
        if goodreviews:
            time.sleep(2)
            self.driver.get(link1)
            customerreviews= self.driver.find_elements_by_css_selector('div[id*="customer_review"]')
            for i in range(1,len(customerreviews)+1):
                ratingcust=self.driver.find_elements_by_xpath('(//div[contains(@id,"customer_review")])['+str(i)+'] /div[@class="a-row"] /a [1]')
                reviewtitle=self.driver.find_elements_by_xpath('(//div[contains(@id,"customer_review")])['+str(i)+'] /div[@class="a-row"] /a [2]')
                if ratingcust or reviewtitle:
                    g_reviews+= reviewtitle[0].text +"("+ ratingcust[0].get_attribute("title") +") \n"
        
        if criticalreviews:
            time.sleep(2)
            #link1= criticalreviews[0].get_attribute("href")
            self.driver.get(link2)
            customerreviews=self.driver.find_elements_by_css_selector('div[id*="customer_review"]')
            #wait2.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR,'div[id*="customer_review"]')))
            for i in range(1,len(customerreviews)+1):
                ratingcust=self.driver.find_elements_by_xpath('(//div[contains(@id,"customer_review")])['+str(i)+'] /div[@class="a-row"] /a [1]')
                reviewtitle=self.driver.find_elements_by_xpath('(//div[contains(@id,"customer_review")])['+str(i)+'] /div[@class="a-row"] /a [2]')
                if ratingcust or reviewtitle:
                    c_reviews+= reviewtitle[0].text +"("+ ratingcust[0].get_attribute("title") +") \n"
        
        return g_reviews,c_reviews

    def get_product_reviews_and_rating(self):
        overallrating= self.driver.find_elements_by_css_selector('span[data-hook="rating-out-of-text"]')
        reviews= self.driver.find_elements_by_css_selector('a[data-hook*="see-all-reviews-link"]')
        rating="NA"
        
        if overallrating:
            rating= overallrating[0].get_attribute("textContent")
        
        good , critical = self.extract_reviews(reviews)
        self.driver.switch_to.window(self.handles[0])
        return (good,critical,rating)
        
               
        
    def exportxlsx(self,path='',name='test1.xlsx'):
        assert name[-5:] ==".xlsx" ,"Invalid sheet type"
        workbook= Workbook()
        sheet=workbook.active
        headers=["Product-Name","Product-Price","Product-Shipping-Time","Product-Shipping-Price","Overview","Features","Rating","Positive Reviews","Critical Reviews","Link","Keyword"]
        r=1
        for c in range(1,len(headers)+1):
            sheet.cell(row=r,column=c).value=headers[c-1]
        
        for product in self.products:
            r+=1
            c=1
            sheet.cell(row=r,column=c).value=product.name
            for i in range(len( product.price)):
                sheet.cell(row=r,column=c+1).value=str(product.price[i])
                if len(product.deliverytime) >i:
                    sheet.cell(row=r,column=c+2).value=str(product.deliverytime[i])
                if len(product.deliveryprice)>i:
                    sheet.cell(row=r,column=c+3).value=str(product.deliveryprice[i])
                sheet.cell(row=r,column=c+4).value= product.overview
                sheet.cell(row=r,column=c+5).value= product.features
                sheet.cell(row=r,column=c+6).value= product.overallrating
                sheet.cell(row=r,column=c+7).value= product.goodreviews
                sheet.cell(row=r,column=c+8).value= product.criticalreviews
                sheet.cell(row=r,column=c+9).value= product.link
                sheet.cell(row=r,column=c+10).value= product.keyword
                if i != len(product.price)-1:
                    
                    r+=1
        workbook.save(path+name)
        
 
        
### Omitted working scrapepage method.

"""    def scrapepage(self,link):
        options,var= self.getoptions()
        if len(options)>=1:
            xpath=self.get_path_option(options[0],var)
            #print(xpath)
            option1=self.driver.find_elements_by_xpath(xpath)
            
        #print(options)
        if len(options)>=1 and option1:
            good , critical, rating= self.get_product_reviews_and_rating(link)
            xpath=self.get_path_option(options[0],var)
            #print(xpath)
            #option1=self.find_elements_by_xpath(xpath)
           # wait.until(EC.presence_of_all_elements_located((By.XPATH,xpath)))
            for op1 in option1:
                self.__click(op1)
               # webdriver.ActionChains(self.driver).move_to_element(op1 ).click().perform()
                try:
                    assert len(options)>1 ,"Execute the below line"
                    option2=self.wait.until(EC.presence_of_all_elements_located((By.XPATH,self.get_path_option(options[1],var))))
                    for op2 in option2:
                        self.__click(op1)
                        self.__click(op2)
                        #webdriver.ActionChains(self.driver).move_to_element(op2).click().perform()
                        #time.sleep(2)
                        name=self.getproductname()
                        duos=self.getproductpriceanddelivery()
                        overview,features,details= self.get_product_overview_and_features()
                        #print(price,delivery)
                        
                        
                        product=Product(name=name,link=link,keyword=self.currentproduct,deliveryoptions=duos,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                        product.deliverybreakdown()
                        self.products.append(product)           
                except:
                    self.__click(op1)
                    time.sleep(1.5)
                    name=self.getproductname()
                    duos=self.getproductpriceanddelivery()
                    overview,features,details= self.get_product_overview_and_features()
                    product=Product(name=name, deliveryoptions=duos,link=link,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
                    product.deliverybreakdown()
                    self.products.append(product)
        else:
            name=self.getproductname()
            duos=self.getproductpriceanddelivery()
            overview,features,details= self.get_product_overview_and_features()
            good , critical, rating= self.get_product_reviews_and_rating(link)
            product=Product(name=name, deliveryoptions=duos,link=link,overview=overview,features=features,goodreviews=good,criticalreviews=critical,rating=rating,details=details)
            product.deliverybreakdown()
            self.products.append(product)
            """
